from flask import Blueprint, request, jsonify, send_file, current_app
from app.models import Project, Branch
from app import db
import pandas as pd
import io
from datetime import datetime
import os
import tempfile

export_bp = Blueprint('export', __name__, url_prefix='/export')

@export_bp.route('/csv')
def export_csv():
    """CSVエクスポート機能"""
    try:
        # 検索条件を取得
        project_code = request.args.get('project_code', '')
        project_name = request.args.get('project_name', '')
        branch_id = request.args.get('branch_id', type=int)
        fiscal_year = request.args.get('fiscal_year', type=int)
        order_probability_min = request.args.get('order_probability_min', type=float)
        order_probability_max = request.args.get('order_probability_max', type=float)
        
        # プロジェクトデータを検索条件に基づいて取得
        query = Project.query.join(Branch)
        
        # 検索条件を適用
        if project_code:
            query = query.filter(Project.project_code.contains(project_code))
        
        if project_name:
            query = query.filter(Project.project_name.contains(project_name))
        
        if branch_id:
            query = query.filter(Project.branch_id == branch_id)
        
        if fiscal_year:
            query = query.filter(Project.fiscal_year == fiscal_year)
        
        if order_probability_min is not None:
            query = query.filter(Project.order_probability >= order_probability_min)
        
        if order_probability_max is not None:
            query = query.filter(Project.order_probability <= order_probability_max)
        
        # データを取得
        projects = query.order_by(Project.created_at.desc()).all()
        
        # データをDataFrame形式に変換
        data = []
        for project in projects:
            data.append({
                'プロジェクトコード': project.project_code,
                'プロジェクト名': project.project_name,
                '支社名': project.branch.branch_name if project.branch else '',
                '支社コード': project.branch.branch_code if project.branch else '',
                '売上の年度': project.fiscal_year,
                '受注角度': f"{project.order_probability_symbol} {int(project.order_probability)}%",
                '受注角度(数値)': int(project.order_probability),
                '売上（契約金）': float(project.revenue),
                '経費（トータル）': float(project.expenses),
                '粗利': project.gross_profit,
                '作成日': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
                '更新日': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else ''
            })
        
        # DataFrameを作成
        df = pd.DataFrame(data)
        
        # CSVファイルを生成
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')  # BOM付きUTF-8でExcelでも文字化けしない
        output.seek(0)
        
        # ファイル名を生成（現在日時を含む）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'projects_export_{timestamp}.csv'
        
        # BytesIOに変換してファイルとして送信
        csv_bytes = io.BytesIO()
        csv_bytes.write(output.getvalue().encode('utf-8-sig'))
        csv_bytes.seek(0)
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'CSV export error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'CSVエクスポート中にエラーが発生しました。'
        }), 500

@export_bp.route('/csv/download-link')
def csv_download_link():
    """CSVダウンロードリンク生成API"""
    try:
        # 検索条件を取得
        conditions = {}
        if request.args.get('project_code'):
            conditions['project_code'] = request.args.get('project_code')
        if request.args.get('project_name'):
            conditions['project_name'] = request.args.get('project_name')
        if request.args.get('branch_id'):
            conditions['branch_id'] = request.args.get('branch_id')
        if request.args.get('fiscal_year'):
            conditions['fiscal_year'] = request.args.get('fiscal_year')
        if request.args.get('order_probability_min'):
            conditions['order_probability_min'] = request.args.get('order_probability_min')
        if request.args.get('order_probability_max'):
            conditions['order_probability_max'] = request.args.get('order_probability_max')
        
        # 対象データ件数を取得
        query = Project.query.join(Branch)
        
        # 検索条件を適用
        if conditions.get('project_code'):
            query = query.filter(Project.project_code.contains(conditions['project_code']))
        if conditions.get('project_name'):
            query = query.filter(Project.project_name.contains(conditions['project_name']))
        if conditions.get('branch_id'):
            query = query.filter(Project.branch_id == int(conditions['branch_id']))
        if conditions.get('fiscal_year'):
            query = query.filter(Project.fiscal_year == int(conditions['fiscal_year']))
        if conditions.get('order_probability_min'):
            query = query.filter(Project.order_probability >= float(conditions['order_probability_min']))
        if conditions.get('order_probability_max'):
            query = query.filter(Project.order_probability <= float(conditions['order_probability_max']))
        
        count = query.count()
        
        # ダウンロードURLを生成
        from urllib.parse import urlencode
        download_url = f"/export/csv?{urlencode(conditions)}"
        
        return jsonify({
            'success': True,
            'download_url': download_url,
            'record_count': count,
            'message': f'{count}件のプロジェクトデータをCSV形式でエクスポートします。'
        })
        
    except Exception as e:
        current_app.logger.error(f'CSV download link generation error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'ダウンロードリンクの生成中にエラーが発生しました。'
        }), 500

@export_bp.route('/excel')
def export_excel():
    """Excelエクスポート機能"""
    try:
        # 検索条件を取得
        project_code = request.args.get('project_code', '')
        project_name = request.args.get('project_name', '')
        branch_id = request.args.get('branch_id', type=int)
        fiscal_year = request.args.get('fiscal_year', type=int)
        order_probability_min = request.args.get('order_probability_min', type=float)
        order_probability_max = request.args.get('order_probability_max', type=float)
        
        # プロジェクトデータを検索条件に基づいて取得
        query = Project.query.join(Branch)
        
        # 検索条件を適用
        if project_code:
            query = query.filter(Project.project_code.contains(project_code))
        
        if project_name:
            query = query.filter(Project.project_name.contains(project_name))
        
        if branch_id:
            query = query.filter(Project.branch_id == branch_id)
        
        if fiscal_year:
            query = query.filter(Project.fiscal_year == fiscal_year)
        
        if order_probability_min is not None:
            query = query.filter(Project.order_probability >= order_probability_min)
        
        if order_probability_max is not None:
            query = query.filter(Project.order_probability <= order_probability_max)
        
        # データを取得
        projects = query.order_by(Project.created_at.desc()).all()
        
        # データをDataFrame形式に変換
        data = []
        for project in projects:
            data.append({
                'プロジェクトコード': project.project_code,
                'プロジェクト名': project.project_name,
                '支社名': project.branch.branch_name if project.branch else '',
                '支社コード': project.branch.branch_code if project.branch else '',
                '売上の年度': project.fiscal_year,
                '受注角度': f"{project.order_probability_symbol} {int(project.order_probability)}%",
                '受注角度(数値)': int(project.order_probability),
                '売上（契約金）': float(project.revenue),
                '経費（トータル）': float(project.expenses),
                '粗利': project.gross_profit,
                '作成日': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
                '更新日': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else ''
            })
        
        # DataFrameを作成
        df = pd.DataFrame(data)
        
        # Excelファイルを生成
        output = io.BytesIO()
        
        # openpyxlを使用してExcelファイルを作成
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='プロジェクト一覧', index=False)
            
            # ワークシートを取得してフォーマットを調整
            worksheet = writer.sheets['プロジェクト一覧']
            
            # 列幅を自動調整
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 最大幅を30に制限
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # ヘッダー行のスタイルを設定
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 数値列の書式設定
            for row in worksheet.iter_rows(min_row=2):
                # 売上（契約金）列
                if len(row) > 7:
                    row[7].number_format = '#,##0'
                # 経費（トータル）列
                if len(row) > 8:
                    row[8].number_format = '#,##0'
                # 粗利列
                if len(row) > 9:
                    row[9].number_format = '#,##0'
        
        output.seek(0)
        
        # ファイル名を生成（現在日時を含む）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'projects_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'Excel export error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'Excelエクスポート中にエラーが発生しました。'
        }), 500

@export_bp.route('/excel/download-link')
def excel_download_link():
    """Excelダウンロードリンク生成API"""
    try:
        # 検索条件を取得
        conditions = {}
        if request.args.get('project_code'):
            conditions['project_code'] = request.args.get('project_code')
        if request.args.get('project_name'):
            conditions['project_name'] = request.args.get('project_name')
        if request.args.get('branch_id'):
            conditions['branch_id'] = request.args.get('branch_id')
        if request.args.get('fiscal_year'):
            conditions['fiscal_year'] = request.args.get('fiscal_year')
        if request.args.get('order_probability_min'):
            conditions['order_probability_min'] = request.args.get('order_probability_min')
        if request.args.get('order_probability_max'):
            conditions['order_probability_max'] = request.args.get('order_probability_max')
        
        # 対象データ件数を取得
        query = Project.query.join(Branch)
        
        # 検索条件を適用
        if conditions.get('project_code'):
            query = query.filter(Project.project_code.contains(conditions['project_code']))
        if conditions.get('project_name'):
            query = query.filter(Project.project_name.contains(conditions['project_name']))
        if conditions.get('branch_id'):
            query = query.filter(Project.branch_id == int(conditions['branch_id']))
        if conditions.get('fiscal_year'):
            query = query.filter(Project.fiscal_year == int(conditions['fiscal_year']))
        if conditions.get('order_probability_min'):
            query = query.filter(Project.order_probability >= float(conditions['order_probability_min']))
        if conditions.get('order_probability_max'):
            query = query.filter(Project.order_probability <= float(conditions['order_probability_max']))
        
        count = query.count()
        
        # ダウンロードURLを生成
        from urllib.parse import urlencode
        download_url = f"/export/excel?{urlencode(conditions)}"
        
        return jsonify({
            'success': True,
            'download_url': download_url,
            'record_count': count,
            'message': f'{count}件のプロジェクトデータをExcel形式でエクスポートします。'
        })
        
    except Exception as e:
        current_app.logger.error(f'Excel download link generation error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'ダウンロードリンクの生成中にエラーが発生しました。'
        }), 500

@export_bp.route('/preview')
def export_preview():
    """エクスポートプレビューAPI"""
    try:
        # 検索条件を取得
        project_code = request.args.get('project_code', '')
        project_name = request.args.get('project_name', '')
        branch_id = request.args.get('branch_id', type=int)
        fiscal_year = request.args.get('fiscal_year', type=int)
        order_probability_min = request.args.get('order_probability_min', type=float)
        order_probability_max = request.args.get('order_probability_max', type=float)
        
        # プロジェクトデータを検索条件に基づいて取得（最大10件のプレビュー）
        query = Project.query.join(Branch)
        
        # 検索条件を適用
        if project_code:
            query = query.filter(Project.project_code.contains(project_code))
        
        if project_name:
            query = query.filter(Project.project_name.contains(project_name))
        
        if branch_id:
            query = query.filter(Project.branch_id == branch_id)
        
        if fiscal_year:
            query = query.filter(Project.fiscal_year == fiscal_year)
        
        if order_probability_min is not None:
            query = query.filter(Project.order_probability >= order_probability_min)
        
        if order_probability_max is not None:
            query = query.filter(Project.order_probability <= order_probability_max)
        
        # 総件数を取得
        total_count = query.count()
        
        # プレビュー用に最大10件を取得
        projects = query.order_by(Project.created_at.desc()).limit(10).all()
        
        # プレビューデータを作成
        preview_data = []
        for project in projects:
            preview_data.append({
                'project_code': project.project_code,
                'project_name': project.project_name,
                'branch_name': project.branch.branch_name if project.branch else '',
                'branch_code': project.branch.branch_code if project.branch else '',
                'fiscal_year': project.fiscal_year,
                'order_probability_symbol': project.order_probability_symbol,
                'order_probability': int(project.order_probability),
                'revenue': float(project.revenue),
                'expenses': float(project.expenses),
                'gross_profit': project.gross_profit,
                'created_at': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
                'updated_at': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else ''
            })
        
        return jsonify({
            'success': True,
            'total_count': total_count,
            'preview_count': len(preview_data),
            'preview_data': preview_data,
            'columns': [
                'プロジェクトコード', 'プロジェクト名', '支社名', '支社コード',
                '売上の年度', '受注角度', '売上（契約金）', '経費（トータル）',
                '粗利', '作成日', '更新日'
            ]
        })
        
    except Exception as e:
        current_app.logger.error(f'Export preview error: {str(e)}')
        return jsonify({
            'success': False,
            'error': 'プレビューの生成中にエラーが発生しました。'
        }), 500