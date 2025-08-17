#!/usr/bin/env python3
"""
Excel Export Integration Test

This script tests the Excel export functionality with actual database data.
"""

import sys
import os
import tempfile
from io import BytesIO
from openpyxl import load_workbook

# Add the app directory to the Python path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'app'))

def test_excel_export_integration():
    """Test Excel export with actual data"""
    try:
        from app import create_app
        from app.models import Project, Branch
        from app import db
        
        # Create Flask app
        app = create_app()
        
        with app.app_context():
            print("Testing Excel export integration...")
            
            # Check if we have any data in the database
            project_count = Project.query.count()
            branch_count = Branch.query.count()
            
            print(f"Database contains {project_count} projects and {branch_count} branches")
            
            if project_count == 0:
                print("No projects found in database. Creating test data...")
                
                # Create a test branch if none exists
                if branch_count == 0:
                    test_branch = Branch(
                        branch_code='TEST',
                        branch_name='テスト支社',
                        is_active=True
                    )
                    db.session.add(test_branch)
                    db.session.commit()
                    print("✓ Test branch created")
                
                # Create a test project
                branch = Branch.query.first()
                test_project = Project(
                    project_code='EXCEL_TEST_001',
                    project_name='Excel Export Test Project',
                    branch_id=branch.id,
                    fiscal_year=2024,
                    order_probability=80.0,
                    revenue=1000000.0,
                    expenses=800000.0
                )
                db.session.add(test_project)
                db.session.commit()
                print("✓ Test project created")
            
            # Test the Excel export endpoint
            with app.test_client() as client:
                print("Testing Excel export endpoint...")
                
                # Test without parameters (export all data)
                response = client.get('/export/excel')
                
                if response.status_code == 200:
                    print("✓ Excel export endpoint responded successfully")
                    
                    # Check content type
                    expected_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    if response.content_type == expected_content_type:
                        print("✓ Correct content type returned")
                    else:
                        print(f"✗ Incorrect content type: {response.content_type}")
                        return False
                    
                    # Check if we got Excel data
                    excel_data = BytesIO(response.data)
                    
                    try:
                        workbook = load_workbook(excel_data)
                        print("✓ Valid Excel file returned")
                        
                        # Check worksheet
                        if 'プロジェクト一覧' in workbook.sheetnames:
                            print("✓ Correct worksheet name found")
                            
                            worksheet = workbook['プロジェクト一覧']
                            
                            # Check headers
                            expected_headers = [
                                'プロジェクトコード', 'プロジェクト名', '支社名', '支社コード',
                                '売上の年度', '受注角度', '受注角度(数値)', '売上（契約金）',
                                '経費（トータル）', '粗利', '作成日', '更新日'
                            ]
                            
                            actual_headers = [cell.value for cell in worksheet[1]]
                            
                            if actual_headers == expected_headers:
                                print("✓ Correct headers found in Excel file")
                            else:
                                print(f"✗ Incorrect headers: {actual_headers}")
                                return False
                            
                            # Check if data rows exist
                            if worksheet.max_row > 1:
                                print(f"✓ Excel file contains {worksheet.max_row - 1} data rows")
                                
                                # Check first data row
                                first_row = [cell.value for cell in worksheet[2]]
                                if first_row[0]:  # Project code should not be empty
                                    print("✓ Data rows contain valid project data")
                                else:
                                    print("✗ Data rows appear to be empty")
                                    return False
                            else:
                                print("✗ No data rows found in Excel file")
                                return False
                            
                        else:
                            print(f"✗ Expected worksheet not found. Available: {workbook.sheetnames}")
                            return False
                        
                        workbook.close()
                        
                    except Exception as e:
                        print(f"✗ Failed to read Excel file: {str(e)}")
                        return False
                    
                else:
                    print(f"✗ Excel export endpoint failed with status {response.status_code}")
                    if response.data:
                        print(f"Response: {response.data.decode('utf-8')[:200]}")
                    return False
                
                # Test with search parameters
                print("Testing Excel export with search parameters...")
                response = client.get('/export/excel?fiscal_year=2024')
                
                if response.status_code == 200:
                    print("✓ Excel export with parameters works")
                else:
                    print(f"✗ Excel export with parameters failed: {response.status_code}")
                    return False
                
                # Test Excel download link endpoint
                print("Testing Excel download link endpoint...")
                response = client.get('/export/excel/download-link')
                
                if response.status_code == 200:
                    data = response.get_json()
                    if data and data.get('success'):
                        print("✓ Excel download link endpoint works")
                        print(f"  Record count: {data.get('record_count', 0)}")
                        print(f"  Download URL: {data.get('download_url', 'N/A')}")
                    else:
                        print(f"✗ Excel download link endpoint returned error: {data}")
                        return False
                else:
                    print(f"✗ Excel download link endpoint failed: {response.status_code}")
                    return False
            
            print("\n🎉 All Excel export integration tests passed!")
            return True
            
    except Exception as e:
        print(f"✗ Integration test failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_export_integration()
    sys.exit(0 if success else 1)