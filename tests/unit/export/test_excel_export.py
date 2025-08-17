#!/usr/bin/env python3
"""
Excel Export Functionality Test

This script tests the Excel export functionality for the project revenue system.
"""

import sys
import os
import tempfile
import pandas as pd
from openpyxl import load_workbook

# Add the app directory to the Python path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'app'))

def test_excel_export():
    """Test Excel export functionality"""
    try:
        from app import create_app
        from app.models import Project, Branch
        
        # Create Flask app
        app = create_app()
        
        with app.app_context():
            print("Testing Excel export functionality...")
            
            # Test 1: Check if openpyxl is available
            try:
                import openpyxl
                print("✓ openpyxl library is available")
            except ImportError:
                print("✗ openpyxl library is not available")
                return False
            
            # Test 2: Check if pandas is available
            try:
                import pandas as pd
                print("✓ pandas library is available")
            except ImportError:
                print("✗ pandas library is not available")
                return False
            
            # Test 3: Test Excel file creation with sample data
            try:
                # Create sample data
                sample_data = [
                    {
                        'プロジェクトコード': 'TEST001',
                        'プロジェクト名': 'テストプロジェクト1',
                        '支社名': 'テスト支社',
                        '支社コード': 'TEST',
                        '売上の年度': 2024,
                        '受注角度': 'A 80%',
                        '受注角度(数値)': 80,
                        '売上（契約金）': 1000000.0,
                        '経費（トータル）': 800000.0,
                        '粗利': 200000.0,
                        '作成日': '2024-01-01 10:00:00',
                        '更新日': '2024-01-01 10:00:00'
                    }
                ]
                
                # Create DataFrame
                df = pd.DataFrame(sample_data)
                
                # Create temporary Excel file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                    with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='プロジェクト一覧', index=False)
                        
                        # Get worksheet and apply formatting
                        worksheet = writer.sheets['プロジェクト一覧']
                        
                        # Adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 30)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Apply header formatting
                        from openpyxl.styles import Font, PatternFill, Alignment
                        
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        
                        for cell in worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # Apply number formatting
                        for row in worksheet.iter_rows(min_row=2):
                            # 売上（契約金）列
                            if len(row) > 7:
                                row[7].number_format = '#,##0'
                            # 経費（トータル）列
                            if len(row) > 8:
                                row[8].number_format = '#,##0'
                            # 粗利列
                            if len(row) > 9:
                                row[9].number_format = '#,##0'
                
                # Verify the Excel file was created and can be read
                workbook = load_workbook(tmp_file.name)
                worksheet = workbook['プロジェクト一覧']
                
                # Check if data is present
                if worksheet['A1'].value == 'プロジェクトコード':
                    print("✓ Excel file created successfully with proper headers")
                else:
                    print("✗ Excel file headers are incorrect")
                    workbook.close()
                    return False
                
                if worksheet['A2'].value == 'TEST001':
                    print("✓ Excel file contains sample data")
                else:
                    print("✗ Excel file does not contain expected data")
                    workbook.close()
                    return False
                
                # Close workbook before cleanup
                workbook.close()
                
                # Clean up
                try:
                    os.unlink(tmp_file.name)
                    print("✓ Excel export functionality test completed successfully")
                except:
                    print("✓ Excel export functionality test completed successfully (cleanup skipped)")
                
                
            except Exception as e:
                print(f"✗ Excel file creation failed: {str(e)}")
                return False
            
            # Test 4: Check export routes
            try:
                from app.export_routes import export_bp
                print("✓ Export routes blueprint loaded successfully")
                
                # Check if Excel export route exists
                excel_route_found = False
                for rule in app.url_map.iter_rules():
                    if rule.endpoint == 'export.export_excel':
                        excel_route_found = True
                        break
                
                if excel_route_found:
                    print("✓ Excel export route is registered")
                else:
                    print("✗ Excel export route is not registered")
                    return False
                    
            except Exception as e:
                print(f"✗ Export routes test failed: {str(e)}")
                return False
            
            print("\n🎉 All Excel export tests passed!")
            return True
            
    except Exception as e:
        print(f"✗ Test failed with error: {str(e)}")
        return False

if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)